
import json, openpyxl, os, glob
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange

class TemplateGenerator():

    def __init__(self, survey, coders=2):
        self._survey = survey
        self._coders = coders

    def generateTemplate(self):
        resps = self.getValidResponses()
        qnames, qtexts = self.getShortAnswerQuestions()
        header = self.getHeader(qnames)
        self.makeRespDir()
        self.makeTemplate(header, resps, qtexts)

    def makeRespDir(self):
        path = os.path.join("surveys", self._survey, "response_coding")
        if not os.path.isdir(path):
            os.makedirs(path)

    def makeTemplate(self, header, resps, qtexts):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        self.addHeaderToSheet(sheet, header, qtexts)
        self.addResponsesToSheet(sheet, resps)

        # Make a sheet for each coder
        self.makeSheets(workbook, sheet)

        # Remove the template sheet
        workbook.remove(sheet)

        # makeCodesSheet(workbook, header, qtexts)

        # Save the workbook
        path = os.path.join("surveys", self._survey, "response_coding", "coding_template.xlsx")
        workbook.save(path)

    def addHeaderToSheet(self, sheet, header, qtexts):
        for i, head in enumerate(header):
            column = chr(65 + i)
            row = "1"
            sheet[column + row] = head
            if not column in "AB":
                sheet[column+row].comment = Comment(qtexts[i-2], "AutoTool")

    def addResponsesToSheet(self, sheet, resps):
        for j, resp in enumerate(resps):
            pid, rid = resp
            row = str(2 + j)
            sheet["A"+row] = int(pid)
            sheet["B"+row] = rid

    def makeSheets(self, workbook, original):
        for coder in range(self._coders):
            copy = workbook.create_sheet(title=f"Coder_{coder+1}")
            self.copySheet(original, copy)

    def copySheet(self, original, copy):
        for row in original.iter_rows(min_row=1, max_row=original.max_row, min_col=1, max_col=original.max_column):
            for cell in row:
                new_cell = copy[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.protection = copy(cell.protection)
                if cell.comment:
                    new_comment = Comment(cell.comment.text, cell.comment.author)
                    new_cell.comment = new_comment

    def getHeader(self, questions):
        return ["JSON", "Response ID"] + questions

    def getShortAnswerQuestions(self):
        path = os.path.join("surveys", self._survey, "questions.json")
        with open(path, "r", encoding="utf-8") as file:
            d = json.load(file)
        qnames = []
        qtexts = []
        for group in d.values():
            for qname, qdata in group.items():
                if qdata["type"] == "short-answer" and \
                    qdata.get("coded") and \
                    not (qdata.get("contains_pii") or qdata.get("ignore")):
                    qnames.append(qname)
                    qtexts.append(qdata["question"])
        return (qnames, qtexts)

    def getValidResponses(self):
        path = self.getSanitized()
        with open(path, "r", encoding="utf-8") as file:
            d = json.load(file)
        return [(key, value["meta"]["ResponseID"]) for key, value in d.items()]

    def getFilePath(self):
        return os.path.join("surveys", self._survey, "files")

    def getSanitized(self):
        path = os.path.join(self.getFilePath(), "*sanitized*.json")
        files = glob.glob(path)
        files = sorted([f.split(os.sep)[-1][:-5] for f in files])
        file = files[-1] + ".json"
        return os.path.join(self.getFilePath(), file)